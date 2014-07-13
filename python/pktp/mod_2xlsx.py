from openpyxl import Workbook  # using openpyxl
import os

def rm_hidden_file(file_list):
    new_list = []
    for file_name in file_list:
        if('~' not in file_name and '.' not in file_name):
            new_list.append(file_name)
    return new_list 


def to_xlsx():
    usr_dir = '/home/pk/usr/'  # should be /usr
    xlsx_name = 'grade.xlsx'

    sections = {}  # key is section number, value is worksheet obj
    id_files = os.listdir(usr_dir) # need to remove hidden file
    id_files = rm_hidden_file(id_files) # remove files contain '.' or '~' in name
    sec_quiz = {} # key is section number, value is a dictionary whose key is quiz name, value is column number
    id_names = {} # key is section number, value is names list in that section

    workbook = Workbook()

    for id_file in id_files:
        f = open(usr_dir + id_file, 'r')
        for line in f:
            item = line.split()
            # check user_id and email
            if(item[0] == 'email'):
                continue
            # class section
            elif(item[0] == 'section'):
                sec_name = item[1]
                if(sec_name not in sections): # if section number not in sections, then create a new sheet for section
                    worksheet = workbook.create_sheet(0)
                    worksheet.title = sec_name
                    sections[sec_name] = worksheet
                    sec_quiz[sec_name] = {}
                else:
                    worksheet = sections[sec_name]
                if(sec_name not in id_names):
                    id_names[sec_name] = [id_file]
                else:
                    id_names[sec_name].append(id_file)
            # fill in grade
            elif(item[0] == 'grade'):
                quiz_name = item[1]
                grade = item[2]
                max_score = item[3]
                if(quiz_name not in sec_quiz[sec_name]):
                    sec_quiz[sec_name][quiz_name] = len(sec_quiz[sec_name]) + 1
                    worksheet.cell(row = 0, column = sec_quiz[sec_name][quiz_name]).value = quiz_name + ' ' + max_score 
                worksheet.cell(row = len(id_names[sec_name]), column = sec_quiz[sec_name][quiz_name]).value = grade
            # error
            else:
                print 'Should have no field called ' + item[0]
                id_file = ''
                f.close()
                break

        if(id_file):
            worksheet.cell(row = len(id_names[sec_name]), column = 0).value = id_file # fill in id
        f.close()
        
    workbook.save(xlsx_name)
