from openpyxl import load_workbook
import os

def from_xlsx():
    out_dir = '/home/pk/out/'  # should be /usr
    xlsx_name = 'grade.xlsx'

    wb = load_workbook(xlsx_name)
    section_names = wb.get_sheet_names()

    print section_names


    for section_name in section_names:
        ws = wb.get_sheet_by_name(section_name)
        end_row = ws.get_highest_row()
        end_column = ws.get_highest_column()
        quiz_names = []
        for quiz_column in range(1, end_column):
            quiz_names.append(str(ws.cell(row = 0, column = quiz_column).value))
        print quiz_names

        for name_row in range(1, end_row):
            file_name = str(ws.cell(row = name_row, column = 0).value)
            print file_name
            f = open(out_dir + file_name, 'w')
            f.write('email ' + file_name + '@binghamton.edu\n')
            f.write('section ' + section_name + '\n')
            for quiz_grade in range(1, len(quiz_names) + 1):
                grade = str(ws.cell(row = name_row, column = quiz_grade).value)
                name_max_score = quiz_names[quiz_grade - 1].split()
                f.write('grade ' + name_max_score[0] + '\t' + grade + '\t' + name_max_score[1] + '\n')
            f.close()



